def vat_work(nalja):
    import pandas as pd
    import os
    from numpy import NaN,NAN,nan
    import openpyxl
    import os
    d_path = "D:\\OneDrive - 오토넷\\문서\\excel Data\\"


    df_machul = pd.read_excel(d_path+"HC41.XLSX",header=0)
    df_machul = df_machul[['업체코드','업체명','총판매']]
    df_machul.업체코드 = df_machul.업체코드.str.replace('00000','')
    df_machul.업체코드 = df_machul.업체코드.str.strip()

    df_list = pd.read_excel(d_path+"거래처.xls",header=9)
    df_list = df_list[['업체(Main)','사업자번호','업체명','대표자명','주소(도로명)','업태','업종','이메일주소']]

    df_vat = pd.merge(df_machul,df_list,left_on='업체코드',right_on='업체(Main)')
    df_vat.사업자번호 = df_vat.사업자번호.str.replace('-','')  #사업자번호에 포함된 "-"삭제
    df_vat.업체코드 = (df_vat.업체코드.str.replace('00000','')).str.strip()
    df_vat = df_vat[(df_vat.업체코드.str.contains('B')) == False]  # 업체코드 B로 시작하는 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('Z')) == False]  # 업체코드 Z로 시작하는 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('0002')) == False]  # 업체코드 0002인 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('106A')) == False]  # 업체코드 106A인 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('M201')) == False]  # 업체코드 일진2부 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('M208')) == False]  # 업체코드 만안공업사 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('M208A')) == False]  # 업체코드 만안공업사 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('9100')) == False]  # 업체코드 신흥3층 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('9000')) == False]  # 업체코드 신용카드판매 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('9998')) == False]  # 업체코드 회사차 수리 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('9999')) == False]  # 업체코드 일반고객 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('6000')) == False]  # 업체코드 온라인판매 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('5007')) == False]  # 업체코드 반석상사 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('6020')) == False]  # 업체코드 서울상사 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('8010')) == False]  # 업체코드 인카 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('10131')) == False]  # 업체코드 금강정비 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('3008')) == False]  # 업체코드 케이투모터스 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('M700')) == False]  # 업체코드 안양학원 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('3011K')) == False]  # 업체코드 GS넥스테이션 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('4100')) == False]  # 업체코드 오토링(박달동) 거래처 삭제
    df_vat = df_vat[(df_vat.업체코드.str.contains('4101')) == False]  # 업체코드 친절카(박달동) 거래처 삭제
    df_vat = df_vat[(df_vat.사업자번호.str.contains('0000000000')) == False]  # 사업자 번호 없는 거래처 삭제
    df_vat.reset_index(inplace=True)
    # df_vat = df_vat.drop("index",1)
    df_vat = df_vat.drop(columns="index")
    # df_vat = df_vat.drop("업체(Main)",1)
    df_vat = df_vat.drop(columns="업체(Main)")

    # df.drop(columns='market')

    # 총판매를 기반으로 부품대,부가세 컬럼 만들어야 한다

    df_vat['(%)'] = 1

    per_093 = ['1006','1010','1062','3023','3023A','2015']   # 군포공업사, 갑지, 신영공업사 , 낙원택시
    per_095 = ['M208','M208A','105','3007','4800']    # 만안공업사, 스피드, 평촌현대, 우일공업사
    per_110 = ['9002']                                # 중앙
    per_130 = ['M502']                         # 태광


    df_vat.loc[df_vat['업체코드'].isin(per_093),'(%)'] = .93
    df_vat.loc[df_vat['업체코드'].isin(per_095),'(%)'] = .95
    df_vat.loc[df_vat['업체코드'].isin(per_110),'(%)'] = 1.1
    df_vat.loc[df_vat['업체코드'].isin(per_130),'(%)'] = 1.3

    df_vat['(합계)'] = round(df_vat['총판매']*df_vat['(%)'])
    df_vat['(금액)'] = round(df_vat['(합계)']/1.1)
    df_vat['(세액)'] = df_vat['(합계)']-df_vat['(금액)']
    df_vat['금액'] = df_vat['(금액)']
    df_vat['세액'] = df_vat['(세액)']

    df_vat.loc[df_vat['총판매']==df_vat['(합계)'],'(비교)'] = 'OK'
    df_vat.loc[df_vat['총판매']!=df_vat['(합계)'],'(비교)'] = ' ** 확인 바람 ** '


    df_vat = df_vat[['업체코드','업체명_x','총판매','(%)','(금액)','(세액)','(합계)','(비교)','금액','세액','사업자번호','업체명_y','대표자명','주소(도로명)','업태','업종','이메일주소']]

    df_vat.to_excel(d_path+"계산서 발행 확인용.xlsx")
    # NoN 값 ""으로 치환
    df_vat['이메일주소']=df_vat['이메일주소'].fillna(value="")
    # NoN 값의 유무 확인
    df_vat.loc[df_vat.isnull()['이메일주소'],:]
    #df_vat
    df_vat.to_excel(d_path+"계산서 발행 확인용.xlsx")

    wb_1 = openpyxl.load_workbook(d_path+"계산서 발행 확인용.xlsx", data_only=False)
    wb_2 = openpyxl.load_workbook(d_path+"세금계산서발행내역.xlsx")

    number =7
    num = 2
    # nal = input("발행일자를 입력 하세요  (ex : 20191231) :      ")
    nal = nalja
    for row in wb_1['Sheet1'].iter_rows(min_row=2):

        old = wb_1['Sheet1']
        aq = '01'
        bq = nal
        cq = '1238163272'
        eq = '(주)오토넷'
        fq = '손희주'
        gq = '경기도 군포시  엘에스로 166번길15-2 (금정동,보람빌딩101호)'
        hq = '도매,소매'
        iq = '자동차부품'
        kq = str(old['L'+str(num)].value)
        mq = str(old['M'+str(num)].value)
        nq = str(old['N'+str(num)].value)
        oq = str(old['O'+str(num)].value)
        pq = str(old['P'+str(num)].value)
        qq = str(old['Q'+str(num)].value)
    # 이메일주소
    #     if str(old['R'+str(num)].value) is None:
        if old['R'+str(num)].value is None:
            rq = ''
        else:
            rq = str(old['R'+str(num)].value)  
        tq = str(old['J'+str(num)].value)
        uq = str(old['K'+str(num)].value)
        wq = nal[6:]
        xq = '부품대'
        zq = '1'
        aaq = str(old['J'+str(num)].value)
        abq = str(old['J'+str(num)].value)
        acq = str(old['K'+str(num)].value)
        bcq = '0'
        bdq = '0'
        beq = '0'
        bfq = str(old['H'+str(num)].value)
        bgq = '02'

        new = wb_2['sheet']


        new['a'+str(number)].value = aq
        new['b'+str(number)].value = bq
        new['c'+str(number)].value = cq
        new['e'+str(number)].value = eq
        new['f'+str(number)].value = fq
        new['g'+str(number)].value = gq
        new['h'+str(number)].value = hq
        new['i'+str(number)].value = iq
        new['k'+str(number)].value = kq
        new['m'+str(number)].value = mq
        new['n'+str(number)].value = nq
        new['o'+str(number)].value = oq
        new['p'+str(number)].value = pq
        new['q'+str(number)].value = qq
        new['r'+str(number)].value = rq
        new['t'+str(number)].value = tq
        new['u'+str(number)].value = uq
        new['w'+str(number)].value = wq
        new['x'+str(number)].value = xq
        new['z'+str(number)].value = zq
        new['aa'+str(number)].value = aaq
        new['ab'+str(number)].value = abq
        new['ac'+str(number)].value = acq
        new['bc'+str(number)].value = bcq
        new['bd'+str(number)].value = bdq
        new['be'+str(number)].value = beq
        new['bf'+str(number)].value = bfq
        new['bg'+str(number)].value = bgq

        number = number + 1
        
        num = num +1
        
    wb_2.save(d_path+'계산서 전송용.xlsx')
    wb_1.close()
    wb_2.close()
    print('작업을 완료 했나봐. 확인 해보셔')
    
    os.startfile("D:\\OneDrive - 오토넷\\문서\\excel Data\\계산서 전송용.xlsx")
    # # 2021년11월 계산서 발행은 이 프로그램으로 했다.

if __name__ == "__main__":
    vat_work()

