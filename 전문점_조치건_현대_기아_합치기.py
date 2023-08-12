import pandas as pd
import ctypes 
import os
import openpyxl
from openpyxl import workbook
from openpyxl.styles import Font,Alignment,Border,Side,Color,PatternFill

현대 = pd.read_excel('C:\\Users\\Jaeri\\Downloads\\현대.xls',header=8,usecols='a,b,e,l,o,p,q,r')
기아 = pd.read_excel('C:\\Users\\Jaeri\\Downloads\\기아.xls',header=8,usecols='a,b,e,l,o,p,q,r')
전문점 = pd.concat([현대,기아])
전문점 = 전문점[['청구일시(순번)','계열','부품번호','부품명','청구','지원센터','가격','청구자']]
전문점.지원센터 = (전문점.지원센터.str.replace('\(주\)','', regex=True)).str.strip()
전문점.지원센터 = (전문점.지원센터.str.replace('주식회사','', regex=True)).str.strip()
전문점 = 전문점.sort_values(by='청구일시(순번)')

전문점.to_excel('C:\\Users\\Jaeri\\Downloads\\전문점 주문 리스트.xlsx')
# 여기까지 pandas로 데이타 가공하고 이제 openpyxl로 양식을 꾸민다

# 장식을 집어넣으면서 코드가 점점 비대해진다. 그저 수작업으로 하던 것을 대체하는 것으로 위안을

# 이제 excel File 꾸미기 작업
wb = openpyxl.load_workbook('C:\\Users\\Jaeri\\Downloads\\전문점 주문 리스트.xlsx')

# 테두리 box 변수
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)

# 매입 sheet 작업
sheet = wb['Sheet1']
sheet["A1"]='NO'
sheet["C1"]='LEP'
sheet["F1"]='수량'
sheet["H1"]='금액'
sheet['J1']='공백 제거 품번'
sheet.column_dimensions['A'].width = 6.25
sheet.column_dimensions['B'].width = 21.25
sheet.column_dimensions['C'].width = 3.75
sheet.column_dimensions['D'].width = 17
sheet.column_dimensions['E'].width = 35.5
sheet.column_dimensions['F'].width = 4.88
sheet.column_dimensions['G'].width = 13
sheet.column_dimensions['H'].width = 15
sheet.column_dimensions['J'].width = 17

m_row = sheet.max_row

for row in range(1,m_row+1):
    sheet.row_dimensions[row].height = 18
      
for row in sheet['A1:J'+str(m_row)]:
    for cell in row:
        cell.border = border_thin

for row in sheet['H2:H'+str(m_row)]:
    for cell in row:
        cell.number_format = '#,###'
        
for row in sheet['A2:A'+str(m_row)]:
    for cell in row:
        cell.value = '=row()-1'
 
font_11 = Font(name='맑은 고딕',size=11,bold=True)

for row in sheet[('D2:D'+str(m_row))]:
    for cell in row:
        cell.font = font_11

for row in sheet[('F2:F'+str(m_row))]:
    for cell in row:
        cell.font = font_11     
           
sheet[('A1')].font = font_11
sheet['A1'].alignment = Alignment(horizontal='center',vertical='center')

# 품번 가운데 정렬은 어색하다.
# for row in sheet[('D2:D'+str(m_row))]:
#     for cell in row:
#         cell.alignment = Alignment(horizontal='center',vertical='center')

for row in sheet[('F2:F'+str(m_row))]:
    for cell in row:
        cell.alignment = Alignment(horizontal='center',vertical='center')     

aa = 1
for row in sheet[('J2:J'+str(m_row))]:
    aa = aa + 1
    for cell in row:
        cell.value = '=substitute(D'+str(aa)+'," ","")'
       
wb.save('C:\\Users\\Jaeri\\Downloads\\전문점 주문 리스트.xlsx')
wb.close()
       

os.startfile("C:\\Users\\Jaeri\\Downloads\\전문점 주문 리스트.xlsx")