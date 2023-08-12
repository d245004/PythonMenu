from openpyxl.xml.constants import MAX_COLUMN
import pandas as pd
from openpyxl import workbook
from openpyxl.styles import Font,Alignment,Border,Side,Color,PatternFill
import openpyxl
import os
import sys 
import xlsxwriter


# path 지정
d_path="C://Users//Jaeri//Downloads//"

# 일반 매출 작업
df_machul = pd.read_excel(d_path+"매출.xls",header=12)
df_machul = df_machul[['업체코드','업체명','총판매']]
df_machul.업체코드 = (df_machul.업체코드.str.replace('00000','')).str.strip()
df_machul = df_machul[(df_machul.업체코드.str.contains('B')) == False]
df_machul = df_machul[(df_machul.업체코드.str.contains('Z')) == False]
df_machul.reset_index(inplace=True)
df_machul=df_machul.drop('index',axis=1)

print("일반매출")

# 매입작업
df_maip = pd.read_excel(d_path+"매입.xls",header=10)
df_maip_1 = df_maip.groupby(['업체(M)','업체명','L'])['금액'].agg(['sum'])
df_maip_1 = df_maip_1.pivot_table(index=['업체(M)','업체명'],columns=['L'])

df_maip_1.to_excel('imsi.xlsx')
df_maip1 = pd.read_excel('imsi.xlsx',header=2)



# df_maip1.rename(columns={df_maip1.columns[2]:'현대',df_maip1.columns[3]:'기아',df_maip1.columns[4]:'기타'},inplace=True)
df_maip1.rename(columns={df_maip1.columns[2]:'현대',df_maip1.columns[3]:'기아'},inplace=True)
df_maip1 = df_maip1.fillna(0)     # 빈데이타 0으로 채운다
# df_maip1['합계']=df_maip1['현대']+df_maip1['기아']       # 필요없는 작업이다. excel에서 작업 할 것임

# 매입작업 중 매출이 있는 건 분리 (ex:대리점건 분리)
df_imsi = df_maip.groupby(['업체(M)','업체명'])['금액'].agg(['sum'])
df_imsi_merge = pd.merge(df_machul,df_imsi,left_on='업체코드',right_on='업체(M)')
df_imsi_merge['차액']=df_imsi_merge['총판매']-df_imsi_merge['sum']
df_imsi_merge.rename(columns={'업체코드':'CODE','업체명':'상호','총판매':'매출','sum':'매입'},inplace=True)

print("매입작업")

df_machul = pd.merge(df_machul,df_imsi,left_on='업체코드',right_on='업체(M)',how='left')
df_machul['차액']=df_machul['총판매']-df_machul['sum']
df_machul = df_machul.fillna(0)   # 빈데이타 0으로 채운다
df_machul.rename(columns={'업체코드':'CODE','업체명':'상호','총판매':'매출','sum':'매입','차액':'차액'},inplace=True)

# 보험 매출 작업
df_boheum = pd.read_excel(d_path+"보험.xls",header=14)
df_boheum.정비업체 = (df_boheum.정비업체.str.replace('-00000','')).str.strip()
df_boheum = df_boheum.groupby(['정비업체','정비업체명'])['총판매금액'].agg(['count','sum'])
df_m = pd.merge(df_boheum,df_machul,left_on='정비업체',right_on='CODE')
df_m = df_m[['CODE','상호','count','sum','매출']]
df_m['total']=df_m['sum']+df_m['매출']
df_m.rename(columns={'count':'건수','sum':'보험매출','매출':'일반매출','total':'합계'},inplace=True)

print("보험매출")

# 작업 한 것을 excel File로 저장한다
with pd.ExcelWriter(d_path+'월결산.xlsx') as writer:  # doctest: +SKIP
     df_machul.to_excel(writer,sheet_name='일반매출')
     df_m.to_excel(writer,sheet_name='보험매출')
     df_maip1.to_excel(writer,sheet_name='매입')
     df_imsi_merge.to_excel(writer,sheet_name='대리점거래')
        
# 여기까지 pandas로 데이타 가공하고 이제 openpyxl로 양식을 꾸민다

# 이제 excel File 꾸미기 작업
wb = openpyxl.load_workbook(d_path+"월결산.xlsx")

# 테두리 box 변수
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)

maip_sheet = wb['매입']
maip_row = maip_sheet.max_row

machul_sheet = wb['일반매출']
machul_row = machul_sheet.max_row

boheum_sheet = wb['보험매출']
boheum_row = boheum_sheet.max_row


# 매입 maip_sheet 작업
maip_sheet = wb['매입']
maip_sheet.freeze_panes ='A2'  # 1열 고정 한다


for idx,row in enumerate(maip_sheet.values):    # 찾아서 삭제하기  **********  중요하다 (처음보는 함수네)    **********
    if row[1] == '0001':
        maip_sheet.delete_rows(idx+1)

for row in maip_sheet.iter_rows(min_row=2,max_col=8):   # 짜증스럽네 모르는 것이 시간만 잡아먹는다. max_col은 1부터 시작 row는 0부터 시작 이라네
    if str(row[1].value) == '0002':
        row[7].value = 'M'
    if str(row[1].value) == '5003':
        row[7].value = 'D'
    if str(row[1].value) == '5005':
        row[7].value = 'D'
    if str(row[1].value) == '5007':
        row[7].value = 'D'
    if str(row[1].value) == '501':
        row[7].value = 'D'
    if str(row[1].value) == '6009':
        row[7].value = 'D'
    if str(row[1].value) == '6015':
        row[7].value = 'D'
    if str(row[1].value) == '6017':
        row[7].value = 'D'
    if str(row[1].value) == '6020':
        row[7].value = 'D'
    if str(row[1].value) == '9999':
        row[7].value = 'C'
    if str(row[1].value) == '6000':
        row[7].value = 'O'
    if str(row[1].value) == '9503':
        row[7].value = 'A'
    if str(row[1].value) == '9998':
        row[7].value = 'Z'
    if str(row[1].value) == 'M612':
        row[7].value = 'D'
    if str(row[1].value) == '7000':
        row[7].value = 'D'

maip_sheet.column_dimensions['H'].hidden=True   # 열 숨김



maip_sheet["A1"]='NO'
maip_sheet['G1']='합계'
maip_sheet.column_dimensions['A'].width = 5
maip_sheet.column_dimensions['b'].width = 6
maip_sheet.column_dimensions['C'].width = 20
maip_sheet.column_dimensions['D'].width = 15
maip_sheet.column_dimensions['E'].width = 15
maip_sheet.column_dimensions['F'].width = 15
maip_sheet.column_dimensions['G'].width = 15



      
for row in maip_sheet['A1:G'+str(maip_row-2)]:
    for cell in row:
        cell.border = border_thin


for row in maip_sheet['D2:G'+str(maip_row+10)]:
    for cell in row:
        cell.number_format = '#,###'
        
for row in maip_sheet['A2:A'+str(maip_row-2)]:
    for cell in row:
        cell.value = '=row()-1'

# 매입에서 현대,기아 구분된 것의 합계를 수식으로 집어넣기
aa = 2
for row in maip_sheet['G2:G'+str(maip_row)]:
    for cell in row:
        cell.value = '=sum(D'+str(aa)+':F'+str(aa)+')'
    aa += 1        


maip_sheet['C'+str(maip_row+1)].value = 'TOTAL'    
maip_sheet['D'+str(maip_row+1)].value = '=sum(D2:D'+str(maip_row)+')'
maip_sheet['E'+str(maip_row+1)].value = '=sum(E2:E'+str(maip_row)+')'
maip_sheet['F'+str(maip_row+1)].value = '=sum(F2:F'+str(maip_row)+')'
maip_sheet['G'+str(maip_row+1)].value = '=sum(G2:G'+str(maip_row)+')'

for row in maip_sheet['C'+str(maip_row+4)+':G'+str(maip_row+6)]:
    for cell in row:
        cell.border = border_thin


maip_sheet['C'+str(maip_row+4)].value = '모비스'
maip_sheet['C'+str(maip_row+5)].value = '전문점'
maip_sheet['C'+str(maip_row+6)].value = '대리점'
maip_sheet['C'+str(maip_row+8)].value = '합계'


# 현대 합계
maip_sheet['D'+str(maip_row+4)].value = '=sumif($H$2:$H$'+str(maip_row)+',"M",$D$2:$D$'+str(maip_row)+')'     #  모비스 합계    
maip_sheet['D'+str(maip_row+5)].value = '=sumif($H$2:$H$'+str(maip_row)+',"",$D$2:$D$'+str(maip_row)+')'      #  전문점 합계
maip_sheet['D'+str(maip_row+6)].value = '=sumif($H$2:$H$'+str(maip_row)+',"D",$D$2:$D$'+str(maip_row)+')'     #  대리점 합계
maip_sheet['D'+str(maip_row+8)].value ='=sum(D'+str(maip_row+4)+':D'+str(maip_row+6)+')'                    #  전체 합계
# 기아 합계
maip_sheet['E'+str(maip_row+4)].value = '=sumif($H$2:$H$'+str(maip_row)+',"M",$E$2:$E$'+str(maip_row)+')'     #  모비스 합계    
maip_sheet['E'+str(maip_row+5)].value = '=sumif($H$2:$H$'+str(maip_row)+',"",$E$2:$E$'+str(maip_row)+')'      #  전문점 합계
maip_sheet['E'+str(maip_row+6)].value = '=sumif($H$2:$H$'+str(maip_row)+',"D",$E$2:$E$'+str(maip_row)+')'     #  대리점 합계
maip_sheet['E'+str(maip_row+8)].value ='=sum(E'+str(maip_row+4)+':E'+str(maip_row+6)+')'                    #  전체 합계
#기타 합계
maip_sheet['F'+str(maip_row+4)].value = '=sumif($H$2:$H$'+str(maip_row)+',"M",$F$2:$F$'+str(maip_row)+')'     #  모비스 합계    
maip_sheet['F'+str(maip_row+5)].value = '=sumif($H$2:$H$'+str(maip_row)+',"",$F$2:$F$'+str(maip_row)+')'      #  전문점 합계
maip_sheet['F'+str(maip_row+6)].value = '=sumif($H$2:$H$'+str(maip_row)+',"D",$F$2:$F$'+str(maip_row)+')'     #  대리점 합계
maip_sheet['F'+str(maip_row+8)].value ='=sum(F'+str(maip_row+4)+':F'+str(maip_row+6)+')'                    #  전체 합계

maip_sheet['G'+str(maip_row+4)].value = '=sum(D'+str(maip_row+4)+':F'+str(maip_row+4)+')'
maip_sheet['G'+str(maip_row+5)].value = '=sum(D'+str(maip_row+5)+':F'+str(maip_row+5)+')'
maip_sheet['G'+str(maip_row+6)].value = '=sum(D'+str(maip_row+6)+':F'+str(maip_row+6)+')'
maip_sheet['G'+str(maip_row+8)].value = '=sum(D'+str(maip_row+8)+':F'+str(maip_row+8)+')'

font_15 = Font(name='맑은 고딕',size=15,bold=True)
cell_sum = maip_sheet[('C'+str(maip_row+2))]
cell_sum.font = font_15


# 보험매출 boheum_sheet 작업
boheum_sheet = wb['보험매출']
boheum_sheet.freeze_panes ='A2'  # 1열 고정 한다
boheum_sheet["A1"]='NO'

# 열 사이즈 조정
boheum_sheet.column_dimensions['C'].width = 30
boheum_sheet.column_dimensions['D'].width = 15
boheum_sheet.column_dimensions['E'].width = 15
boheum_sheet.column_dimensions['F'].width = 15
boheum_sheet.column_dimensions['G'].width = 15



for row in boheum_sheet['A1:G'+str(boheum_row)]:
    for cell in row:
        cell.border = border_thin

for row in boheum_sheet['E2:G'+str(boheum_row+10)]:
    for cell in row:
        cell.number_format = '#,###'
        
for row in boheum_sheet['A2:A'+str(boheum_row)]:
    for cell in row:
        cell.value = '=row()-1'

aa = 2
for row in boheum_sheet['G2:G'+str(boheum_row)]:
    for cell in row:
        cell.value = '=E'+str(aa)+'+F'+str(aa) 
    aa += 1       

bb = 2
for row in boheum_sheet['F2:F'+str(machul_row)]:

    for cell in row:
        cell.value = '=iferror(vlookup(B'+str(bb)+',일반매출!$B$2:$D$'+str(machul_row)+',3,0),"")'
    bb += 1


boheum_sheet['C'+str(boheum_row+2)].value = 'TOTAL'    
boheum_sheet['D'+str(boheum_row+2)].value = '=sum(D2:D'+str(boheum_row)+')'
boheum_sheet['E'+str(boheum_row+2)].value = '=sum(E2:E'+str(boheum_row)+')'
boheum_sheet['F'+str(boheum_row+2)].value = '=sum(F2:F'+str(boheum_row)+')'
boheum_sheet['G'+str(boheum_row+2)].value = '=sum(G2:G'+str(boheum_row)+')'

font_15 = Font(name='맑은 고딕',size=15,bold=True)
cell_sum = boheum_sheet[('C'+str(boheum_row+2))]
cell_sum.font = font_15


# 일반매출 machul_sheet 작업
machul_sheet = wb['일반매출']

machul_sheet.freeze_panes ='A2'  # 1열 고정 한다


for row in machul_sheet.iter_rows(min_row=2,max_col=7):   # 짜증스럽네 모르는 것이 시간만 잡아먹는다. max_col은 1부터 시작 row는 0부터 시작 이라네
    if str(row[1].value) == '0002':
        row[6].value = 'M'
    if str(row[1].value) == '5003':
        row[6].value = 'D'
    if str(row[1].value) == '5005':
        row[6].value = 'D'
    if str(row[1].value) == '5007':
        row[6].value = 'D'
    if str(row[1].value) == '501':
        row[6].value = 'D'
    if str(row[1].value) == '6009':
        row[6].value = 'D'
    if str(row[1].value) == '6015':
        row[6].value = 'D'
    if str(row[1].value) == '6017':
        row[6].value = 'D'
    if str(row[1].value) == '6020':
        row[6].value = 'D'
    if str(row[1].value) == '9999':
        row[6].value = 'C'
    if str(row[1].value) == '6000':
        row[6].value = 'O'
    if str(row[1].value) == '9503':
        row[6].value = 'A'
    if str(row[1].value) == '9998':
        row[6].value = 'Z'
    if str(row[1].value) == 'M612':
        row[6].value = 'D'
    if str(row[1].value) == '7000':
        row[6].value = 'D'
    if str(row[1].value) == '5002':
        row[6].value = 'D'
    if 'Z' in str(row[1].value):
        row[6].value = 'A'


machul_sheet.column_dimensions['G'].hidden=True   # 열 숨김




machul_sheet["A1"]='NO'

machul_sheet.column_dimensions['B'].width = 10
machul_sheet.column_dimensions['C'].width = 30
machul_sheet.column_dimensions['D'].width = 20
machul_sheet.column_dimensions['E'].width = 15
machul_sheet.column_dimensions['F'].width = 15


for row in machul_sheet['D2:F'+str(machul_row+15)]:
    for cell in row:
        cell.number_format = '#,###'

for row in machul_sheet['A2:A'+str(machul_row)]:
    for cell in row:
        cell.value = '=row()-1'

for row in machul_sheet['A1:F'+str(machul_row)]:
    for cell in row:
        cell.border = border_thin

aa = 2
for row in machul_sheet['F2:F'+str(machul_row)]:
    for cell in row:
        if machul_sheet['G'+str(aa)].value == 'D':
            cell.value = '=sum(D'+str(aa)+'-E'+str(aa)+')'
        else:
            machul_sheet['F'+str(aa)].value = ''
    aa += 1        

bb = 2
for row in machul_sheet['E2:E'+str(machul_row)]:

    for cell in row:
        cell.value = '=iferror(vlookup(B'+str(bb)+',매입!$B$2:$G$'+str(maip_row)+',6,0),"")'
    bb += 1




machul_sheet['C'+str(machul_row+1)].value = 'TOTAL'    
machul_sheet['D'+str(machul_row+1)].value = '=sum(D2:D'+str(machul_row)+')'

for row in machul_sheet['C'+str(machul_row+4)+':F'+str(machul_row+7)]:
    for cell in row:
        cell.border = border_thin


machul_sheet['C'+str(machul_row+4)].value = '모비스'
machul_sheet['C'+str(machul_row+5)].value = '일반'
machul_sheet['C'+str(machul_row+6)].value = '현금'
machul_sheet['C'+str(machul_row+7)].value = '보험'
machul_sheet['C'+str(machul_row+9)].value = '합계'

machul_sheet['D'+str(machul_row+4)].value = '=sumif($G$2:$G$'+str(machul_row)+',"M",$D$2:$D$'+str(machul_row)+')'     #  모비스 합계    
machul_sheet['D'+str(machul_row+5)].value = '=sumif($G$2:$G$'+str(machul_row)+',"",$D$2:$D$'+str(machul_row)+')'      #  일반 합계
machul_sheet['D'+str(machul_row+6)].value = '=sumif($G$2:$G$'+str(machul_row)+',"C",$D$2:$D$'+str(machul_row)+')'     #  현금 합계
machul_sheet['D'+str(machul_row+7)].value = '=보험매출!E'+str(boheum_row+2)
machul_sheet['E'+str(machul_row+4)].value = '=sumif($G$2:$G$'+str(machul_row)+',"M",$E$2:$E$'+str(machul_row)+')'     #  대리점 매입 합계
machul_sheet['E'+str(machul_row+5)].value = '=sumif($G$2:$G$'+str(machul_row)+',"",$E$2:$E$'+str(machul_row)+')'     #  대리점 매입 합계
machul_sheet['E'+str(machul_row+6)].value = '=sumif($G$2:$G$'+str(machul_row)+',"C",$E$2:$E$'+str(machul_row)+')'     #  대리점 매입 합계
machul_sheet['D'+str(machul_row+9)].value ='=sum(D'+str(machul_row+4)+':D'+str(machul_row+7)+')'                    #  전체 합계

for row in machul_sheet['C'+str(machul_row+11)+':F'+str(machul_row+14)]:
    for cell in row:
        cell.border = border_thin



machul_sheet['C'+str(machul_row+11)].value = '대리점'
machul_sheet['C'+str(machul_row+12)].value = '온라인'
machul_sheet['C'+str(machul_row+13)].value = '오토커머스'
machul_sheet['C'+str(machul_row+14)].value = '회사차 수리'
machul_sheet['D'+str(machul_row+11)].value = '=sumif($G$2:$G$'+str(machul_row)+',"D",$D$2:$D$'+str(machul_row)+')'     #  대리점 매출 합계
machul_sheet['D'+str(machul_row+12)].value = '=sumif($G$2:$G$'+str(machul_row)+',"O",$D$2:$D$'+str(machul_row)+')'     #  온라인 합계
machul_sheet['D'+str(machul_row+13)].value = '=sumif($G$2:$G$'+str(machul_row)+',"A",$D$2:$D$'+str(machul_row)+')'     #  오토커머스 합계
machul_sheet['D'+str(machul_row+14)].value ='=sumif($G$2:$G$'+str(machul_row)+',"Z",$D$2:$D$'+str(machul_row)+')'     #  회사차수리 합계
machul_sheet['E'+str(machul_row+11)].value = '=sumif($G$2:$G$'+str(machul_row)+',"D",$E$2:$E$'+str(machul_row)+')'     #  대리점 매입 합계
machul_sheet['E'+str(machul_row+12)].value = '=sumif($G$2:$G$'+str(machul_row)+',"O",$E$2:$E$'+str(machul_row)+')'     #  대리점 매입 합계
machul_sheet['E'+str(machul_row+13)].value = '=sumif($G$2:$G$'+str(machul_row)+',"A",$E$2:$E$'+str(machul_row)+')'     #  대리점 매입 합계
machul_sheet['E'+str(machul_row+14)].value = '=sumif($G$2:$G$'+str(machul_row)+',"Z",$E$2:$E$'+str(machul_row)+')'     #  대리점 매입 합계
machul_sheet["F"+str(machul_row+11)].value = '=(D'+str(machul_row+7)+'-E'+str(machul_row+7)+')'                        #  대리점 차액




font_15 = Font(name='맑은 고딕',size=15,bold=True)
cell_sum = machul_sheet[('C'+str(machul_row+1))]
cell_sum.font = font_15
cell_total = machul_sheet[('C'+str(machul_row+9))]
cell_total1 = machul_sheet[('D'+str(machul_row+9))]
cell_total.font = font_15
cell_total1.font = font_15
# 대리점거래 maip_sheet 작업
maip_sheet = wb['대리점거래']
maip_sheet.freeze_panes ='A2'  # 1열 고정 한다


for idx,row in enumerate(maip_sheet.values):    # 찾아서 삭제하기  **********  중요하다 (처음보는 함수네)    **********
    if row[1] == '0002':
        maip_sheet.delete_rows(idx+1)
    if row[1] == '9999':
        maip_sheet.delete_rows(idx+1)


maip_sheet["A1"]='NO'

maip_sheet.column_dimensions['B'].width = 10
maip_sheet.column_dimensions['C'].width = 30
maip_sheet.column_dimensions['D'].width = 15
maip_sheet.column_dimensions['E'].width = 15
maip_sheet.column_dimensions['F'].width = 15

s_row = maip_sheet.max_row

for row in maip_sheet['D2:F'+str(s_row+10)]:
    for cell in row:
        cell.number_format = '#,###'
        
for row in maip_sheet['A2:A'+str(s_row-2)]:   #  삭제된 행 갯수 만큼 빼준다
    for cell in row:
        cell.value = '=row()-1'
for row in maip_sheet['A1:F'+str(s_row-2)]:
    for cell in row:
        cell.border = border_thin

aa = 2
for row in maip_sheet['F2:F'+str(s_row-2)]:
    for cell in row:
        cell.value = '=D'+str(aa)+'-E'+str(aa)
    aa += 1

bb = 2
for row in maip_sheet['D2:D'+str(s_row-2)]:
    for cell in row:
        cell.value = '=iferror(vlookup(B'+str(bb)+',일반매출!$B$2:$D$'+str(machul_row)+',3,0),"")'
    bb += 1

cc = 2
for row in maip_sheet['E2:E'+str(s_row-2)]:
    for cell in row:
        cell.value = '=iferror(vlookup(B'+str(cc)+',매입!$B$2:$G$'+str(maip_row)+',6,0),"")'
    cc += 1



        
maip_sheet['C'+str(s_row+1)].value = 'TOTAL'    
maip_sheet['D'+str(s_row+1)].value = '=sum(D2:D'+str(s_row)+')'
maip_sheet['E'+str(s_row+1)].value = '=sum(E2:E'+str(s_row)+')'
maip_sheet['F'+str(s_row+1)].value = '=sum(F2:F'+str(s_row)+')'



font_15 = Font(name='맑은 고딕',size=15,bold=True)
cell_sum = maip_sheet[('C'+str(s_row+1))]
cell_sum.font = font_15



# 작업 내용 저장
wb.save(d_path+"월결산.xlsx")
wb.close()

print("작업종료")

print("모든 작업을 종료 합니다")

try:
    os.startfile(d_path+"월결산.xlsx")
except:
    pass

