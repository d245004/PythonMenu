import openpyxl
import pandas as pd
import os

df_machul = pd.read_excel(
    "C:\\Users\\Jaeri\\Downloads\\song_sale.xls", header=10)
df_machul.rename(columns={'업체(M)': '업체'}, inplace=True)
df_machul = df_machul[(df_machul.업체.str.contains('Z')) == True]
df_machul = df_machul.sort_values(
    by=['일자', '업체', '증표번호'])  # 여러열 오름 차순으로 정리 하기
df_machul = df_machul[df_machul.특이사항 != 'ABC']
df_machul.to_excel("C://Users//Jaeri//Downloads//song_sale.xlsx")

wb1 = openpyxl.load_workbook(
    "C://Users//Jaeri//Downloads//song_sale.xlsx")
wb2 = openpyxl.load_workbook(
    "C://Users//Jaeri//Downloads//eCount 자료입력 Data.xlsx")
ws1 = wb1.active
ws2 = wb2['구매입력']
ws3 = wb2['품목 등록']
ws4 = wb2['판매입력']

aa = 2
for row in ws1.iter_rows(min_row=2):
    nal = str(row[2].value)
    nalnal = nal.replace("-", "")
    part = row[7].value
    partname = row[8].value
    qty = row[9].value
    dan_price = row[11].value
    hap_price = row[12].value
    sangho = row[3].value                       # 업체코드로 거래처 결정한다 (신 버전)
    sangho = sangho.replace("Z", "0")            # 업체 코드 일치 시키기

    ws2['A'+str(aa)].value = nalnal
    ws2['G'+str(aa)].value = '22'
    ws2['J'+str(aa)].value = part
    ws2['M'+str(aa)].value = qty
    ws2['N'+str(aa)].value = dan_price
    ws2['P'+str(aa)].value = hap_price
    ws2['C'+str(aa)].value = '00001'

    ws3['A'+str(aa)].value = part
    ws3['B'+str(aa)].value = partname
    ws3['J'+str(aa)].value = dan_price
    ws3['L'+str(aa)].value = dan_price

    ws4['A'+str(aa)].value = nalnal
    ws4['C'+str(aa)].value = sangho[0:5]
    ws4['G'+str(aa)].value = '12'
    ws4['K'+str(aa)].value = part
    ws4['M'+str(aa)].value = qty
    ws4['N'+str(aa)].value = dan_price
    ws4['P'+str(aa)].value = hap_price

    aa += 1
wb1.close()
wb2.save("C://Users//Jaeri//Downloads//eCount_자료.xlsx")
wb2.close()
print("작업종료")

print("모든 작업을 종료 합니다")

os.startfile("C://Users//Jaeri//Downloads//eCount_자료.xlsx")
