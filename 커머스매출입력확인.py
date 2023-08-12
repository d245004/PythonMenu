# 오토넷 매출 커머스에 입력 된것 정상유무 판단하기
import pandas as pd
import os
import openpyxl
from openpyxl import workbook
from openpyxl.styles import Font,Alignment,Border,Side,Color,PatternFill
import warnings
warnings.simplefilter("ignore")

aa = pd.read_excel("C:\\Users\\Jaeri\\Downloads\\song_sale.xlsx", header=0)
aa['오토넷'] = aa['금액']
aa['업체코드'] = aa.업체.str.replace('Z','0')
aa = aa[['업체코드', '업체명', '오토넷']]
aa = pd.pivot_table(aa, index=['업체코드', '업체명'], values='오토넷', aggfunc='sum')
aa.reset_index

bb = pd.read_excel("C:\\Users\\Jaeri\\Downloads\\커머스.xlsx", header=1)
bb['거래처코드'] = bb['거래처코드'].astype(str).str.zfill(7)
bb['거래처코드'] = bb.거래처코드.str.replace('\.0','')
bb['커머스'] = bb['합계']
bb['업체코드'] = bb['거래처코드']
bb = bb[['업체코드', '거래처명', '커머스']]
bb['업체명'] = bb['거래처명']
bb = pd.pivot_table(bb, index=['업체코드', '업체명'], values='커머스', aggfunc='sum')

chk = pd.merge(aa, bb, on=['업체코드','업체명'], how='left')
chk['확인'] = chk['오토넷'] - chk['커머스']
chk['OK'] = chk.apply(lambda x: 'Good' if x['오토넷'] == x['커머스'] else 'Bad', axis=1)
chk.to_excel("C:\\Users\\Jaeri\\Downloads\\매출비교.xlsx")

# 이제 excel File 꾸미기 작업
wb = openpyxl.load_workbook('C:\\Users\\Jaeri\\Downloads\\매출비교.xlsx')

# 테두리 box 변수
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)

# 매입 sheet 작업
sheet = wb['Sheet1']
sheet.column_dimensions['A'].width = 15
sheet.column_dimensions['B'].width = 21.25
sheet.column_dimensions['C'].width = 15
sheet.column_dimensions['D'].width = 15
sheet.column_dimensions['E'].width = 10
sheet.column_dimensions['F'].width = 10

m_row = sheet.max_row

for row in range(1,m_row+1):
    sheet.row_dimensions[row].height = 18
      
for row in sheet['A1:F'+str(m_row)]:
    for cell in row:
        cell.border = border_thin

for row in sheet['C2:D'+str(m_row)]:
    for cell in row:
        cell.number_format = '#,###'
        
wb.save('C:\\Users\\Jaeri\\Downloads\\매출비교.xlsx')
wb.close()
       
os.startfile("C:\\Users\\Jaeri\\Downloads\\매출비교.xlsx")